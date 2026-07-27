"""
报告导出服务 — 支持 Excel (.xlsx) 与 PDF 格式
"""
import io
import logging
from datetime import datetime, timezone
from typing import Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl 未安装，Excel 导出功能不可用")


def export_assessment_excel(
    summary: str = "",
    disease_risk_percent: float = 0.0,
    severity: str = "",
    device: str = "",
    probabilities: Optional[dict] = None,
    risk_tier: str = "低风险",
    responsible_person: str = "",
    deadline_days: str = "",
    treatment_plan: Optional[dict] = None,
) -> Optional[io.BytesIO]:
    """生成评估报告的 Excel 文件，返回 BytesIO 对象"""
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl 未安装，无法导出 Excel")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "评估报告"

    # ── 样式定义 ──
    header_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    section_font = Font(name="微软雅黑", size=11, bold=True, color="1B5E20")
    section_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    label_font = Font(name="微软雅黑", size=10, bold=True)
    value_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def write_header(ws, title: str, row: int = 1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 36
        return row + 1

    def write_section(ws, title: str, row: int):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        ws.row_dimensions[row].height = 26
        return row + 1

    def write_kv(ws, label: str, value: str, row: int):
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = label_font
        c1.border = thin_border
        c2 = ws.cell(row=row, column=2, value=value)
        c2.font = value_font
        c2.border = thin_border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 22

    def write_value(ws, value: str, row: int, col: int = 1):
        c = ws.cell(row=row, column=col, value=value)
        c.font = value_font
        c.border = thin_border
        ws.row_dimensions[row].height = 22

    r = 1
    # ── 标题 ──
    r = write_header(ws, "农作物病虫害 AI 识别评估报告", r)
    r += 1

    # ── 基本信息 ──
    r = write_section(ws, "基本信息", r)
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    row_data = [
        ("生成时间", now_str),
        ("设备", device or "-"),
        ("严重程度", severity or "-"),
        ("风险评分", f"{disease_risk_percent:.1f}%"),
        ("风险等级", risk_tier),
    ]
    if responsible_person:
        row_data.append(("责任人", responsible_person))
    if deadline_days:
        row_data.append(("处置期限", f"{deadline_days} 天"))
    for k, v in row_data:
        write_kv(ws, k, v, r)
        r += 1
    r += 1

    # ── 概率分布 ──
    if probabilities:
        r = write_section(ws, "严重程度概率分布", r)
        for name, value in probabilities.items():
            pct = f"{float(value.rstrip('%')):.1f}%" if isinstance(value, str) and '%' in value else f"{float(value) * 100:.1f}%"
            write_kv(ws, name, pct, r)
            r += 1
        r += 1

    # ── 摘要 ──
    if summary:
        r = write_section(ws, "评估摘要", r)
        write_value(ws, summary, r)
        r += 1
        r += 1

    # ── 防治方案 ──
    if treatment_plan:
        r = write_section(ws, "防治方案", r)
        plan_rows = []
        if treatment_plan.get("urgency"):
            plan_rows.append(("紧急程度", treatment_plan["urgency"]))
        if treatment_plan.get("estimated_cost"):
            plan_rows.append(("预计成本", treatment_plan["estimated_cost"]))
        if treatment_plan.get("treatment_duration"):
            plan_rows.append(("周期", treatment_plan["treatment_duration"]))
        if treatment_plan.get("symptoms"):
            plan_rows.append(("症状", "; ".join(treatment_plan["symptoms"][:5])))
        if treatment_plan.get("causes"):
            plan_rows.append(("发病原因", "; ".join(treatment_plan["causes"][:5])))
        if treatment_plan.get("prevention"):
            plan_rows.append(("预防措施", "; ".join(treatment_plan["prevention"][:5])))
        if treatment_plan.get("chemical_treatment"):
            plan_rows.append(("化学防治", "; ".join(treatment_plan["chemical_treatment"][:5])))
        if treatment_plan.get("organic_treatment"):
            plan_rows.append(("绿色防治", "; ".join(treatment_plan["organic_treatment"][:5])))
        for k, v in plan_rows:
            write_kv(ws, k, v, r)
            r += 1
        r += 1

    # ── 列宽 ──
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
